#find josephus out of n people when kill one out of every q people
import csv
from openpyxl import load_workbook
import zipfile

EMPTY = ['',None]

ZIPPATH = 'ADcarry_xlsx.zip'
CSVPATH = 'ADcarry/ADcarry.csv'
EXCELPATH = 'ADcarry/ADcarry.xlsx'

step=3
BeginPerson='卡莎'

class person:

    def __init__(self,name = 'none',id = 'none'):
        self.name = name
        self.id = id

    def change_name_to(self,new_name):
        self.name = new_name
        return None

    def change_id_to(self,new_id):
        self.id = new_id
        return None

#基类，功能：清空空白项
class Reader:
    def __init__(self):
        pass

    def clean_blank(self, list_temp):#清理列表中的空行
        while 1:
            i = 0
            for x in list_temp:
                if (x.name in EMPTY)|(x.id in EMPTY) :# 等价于if (x.name == '') | (x.id == '')|(x.name == None) | (x.id == None):
                    i = 1
                    list_temp.remove(x)
            if i == 0:
                break
        return None

class ExcelReader(Reader):
    fdir = ''
    def __init__(self,file_dir):
        self.fdir = file_dir

    def read_excel(self):
        wb = load_workbook(filename=self.fdir)
        ws = wb.active
        list = []

        for row in ws.iter_rows():  # 循环读取每一行
            a = person(name=row[0].value, id=row[1].value)  # 每一行的第一列为名字，第二列为学号，创建对象进行存储
            list.append(a)

        self.clean_blank(list)

        return list

class CsvReader(Reader):
    fdir
    def __init__(self, file_dir):
        self.fdir = file_dir

    def read_csv(self):
        with open(self.fdir, newline='', encoding='utf-8')as f:
            f_csv = csv.reader(f)
            list = []

            for row in f_csv:  # 循环读取每行
                a = person(name=row[0], id=row[1])  # 创建对象存储数据
                list.append(a)

        self.clean_blank(list)

        return list

class ZipReader(ExcelReader,CsvReader):
    fdir = ''
    def __init__(self,file_dir):
        self.fdir = file_dir

    def read_zip(self):
        with zipfile.ZipFile(self.fdir) as zfiles:
            namelist = zfiles.namelist()
            self.fdir = namelist[0]  # 获取压缩包内文件名
            list = []

            zfiles.extractall()  # 解压

            # 根据文件类型选择不同方法打开文件
            if self.fdir.endswith('csv'):
                list = self.read_csv()
            elif self.fdir.endswith('xls') | self.fdir.endswith('xlsx'):
                list = self.read_excel()
            else:
                raise Exception('undefined file type')

        return list

def get_death_order(input_list, step, start_person ):   #需要将其写成类
    death_list = []
    list_copy = input_list.copy()
    rem = len(list_copy)

################输入检查，起始位置是否在列表中
    start_point = -1
    for x in list_copy:
        if x.name == start_person:
            start_point = list_copy.index(x)
            print('\033[0;32;40m\t start person founded \033[0m')

    if start_point == -1:
        start_point = 0             #若列表中没有找起始值，则将起始值置为0
        print('\033[0;31;40m\t start person not founded, the start person has been set to the first one \033[0m')
#################
    assert start_point >= 0
    for counter in range(len(list_copy)):
        assert step != 0
        '''判断输入的步长的正负-----------------------'''
        if step > 0:
            out_point = (start_point -1 + step )%rem

        if step < 0:
            out_point = (start_point + 1 + step) % rem
        '''-----------------------------------------------'''

        death_list.append(list_copy[out_point])
        del list_copy[out_point]

        start_point = out_point
        rem = rem -1

    return death_list

class josephus():

    def __init__(self, ipt_list_, step_, start_person_):
        self.list = ipt_list_
        self.step = step_
        self.start_person



    def __iter__(self):

        return

    def __next__(self):

        return

if __name__ == '__main__':
    #调用不同的Reader进行文件读取
    # list = ExcelReader(excel_fdir).read_excel()
    list = CsvReader(csv_fdir).read_csv()
    # list = ZipReader(zip_fdir).read_zip()

    #文件读取结果打印
    for x in list:
        print("name:",x.name,"id:",x.id)

#josephus排序并打印
    DeathList = get_death_order(list,step,BeginPerson)
    print('\033[0;32;40m\t the death order is: \033[0m')
    for y in DeathList:
        print("name:",y.name,"id:",y.id)