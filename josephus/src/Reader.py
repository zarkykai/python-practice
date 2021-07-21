import csv
from openpyxl import load_workbook
import zipfile
from abc import ABC, abstractmethod
from josephus.src.Person import Person

EMPTY = ['', None]


class Reader(ABC):
    def __init__(self):
        pass

    @abstractmethod
    def read(self):
        pass

    @staticmethod
    def clean_blank(list_temp):        # 清理列表中的空白项
        while 1:
            i = 0
            for x in list_temp:
                if (x.name in EMPTY) | (x.id in EMPTY):
                    i = 1
                    list_temp.remove(x)     # 若检查到空白项则移除此项
            if i == 0:
                break
        return None


class ExcelReader(Reader):

    def __init__(self, file_dir):
        super(Reader, self).__init__()
        self.fdir = file_dir

    def read(self):
        wb = load_workbook(filename=self.fdir)
        ws = wb.active
        list_ = []

        for row in ws.iter_rows():  # 循环读取每一行
            a = Person(name=row[0].value, id=row[1].value)  # 每一行的第一列为名字，第二列为学号，创建对象进行存储
            a = Person.Person(name=row[0].value, id=row[1].value)  # 每一行的第一列为名字，第二列为学号，创建对象进行存储
            list_.append(a)

        self.clean_blank(list_)

        return list_


class CsvReader(Reader):
    def __init__(self, file_dir):
        super(Reader, self).__init__()
        self.fdir = file_dir

    def read(self):
        with open(self.fdir, newline='', encoding='utf-8')as f:
            f_csv = csv.reader(f)
            list_ = []

            for row in f_csv:  # 循环读取每行
                a = Person(name=row[0], id=row[1])  # 创建对象存储数据
                a = Person.Person(name=row[0], id=row[1])  # 创建对象存储数据
                list_.append(a)

        self.clean_blank(list_)

        return list_


class ZipReader(ExcelReader, CsvReader):
    def __init__(self, file_dir, target_file):
        super(Reader, self).__init__()
        self.__fdir = file_dir
        self.__target_file = target_file

    def read(self):
        with zipfile.ZipFile(self.__fdir) as zfiles:
            name_list = zfiles.namelist()
            print(name_list)
            assert self.__target_file in name_list

            list_ = []

            zfiles.extractall()  # 解压

            # 根据文件类型选择不同方法打开文件
            if self.__target_file.endswith('csv'):
                list_ = CsvReader(self.__target_file).read()
            elif self.__target_file.endswith('xls') | self.__target_file.endswith('xlsx'):
                list_ = ExcelReader(self.__target_file).read()
            else:
                raise Exception('unsupported file type')

            self.clean_blank(list_)

        return list_
